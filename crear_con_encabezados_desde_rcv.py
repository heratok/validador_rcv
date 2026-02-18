import json
import os
import pandas as pd
from openpyxl import load_workbook

# --- CONFIGURACION ---
BASE_DIR = "."
ARCHIVO_ENCABEZADOS_JSON = "encabezados.json"
DATA_START_ROW = 4  # 1-based; datos empiezan en fila 4


def detectar_engine(archivo):
    ext = os.path.splitext(archivo)[1].lower()
    if ext == ".xlsb":
        return "pyxlsb"
    if ext in [".xlsx", ".xls"]:
        return "openpyxl"
    return None


def leer_encabezados_json(archivo):
    if not os.path.exists(archivo):
        raise FileNotFoundError(
            f"No se encontro el archivo de encabezados: {archivo}. "
            "Genere el JSON con guardar_encabezados.py"
        )
    with open(archivo, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list) or not data:
        raise ValueError("El JSON de encabezados debe ser una lista no vacia")
    return data


def procesar_archivo(ruta_archivo, encabezados, data_start_row=DATA_START_ROW):
    engine = detectar_engine(ruta_archivo)
    if not engine:
        print(f"Saltando (formato no soportado): {ruta_archivo}")
        return

    # Leer datos desde fila 4 (skiprows=3), sin encabezados
    if engine == "openpyxl":
        # Usar openpyxl para conservar el valor mostrado en Excel (formato incluido)
        def infer_decimals(fmt):
            if not fmt or fmt == "General":
                return None
            # Quitar secciones y literales
            fmt = fmt.split(";")[0]
            fmt = fmt.replace('"', "")
            if "." in fmt:
                dec_part = fmt.split(".", 1)[1]
                dec_part = "".join(ch for ch in dec_part if ch in "0#")
                return len(dec_part)
            return 0

        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active
        data_rows = []
        max_col = ws.max_column
        
        # Detectar la última fila con datos basándose en la PRIMERA COLUMNA (consecutivo)
        # Un registro válido = tiene consecutivo (primera columna)
        max_row = ws.max_row
        while max_row >= data_start_row:
            # Revisar si la primera columna tiene un valor
            cell = ws.cell(row=max_row, column=1)
            if cell.value is not None:
                break
            max_row -= 1
        
        # Si no encontró datos, usar DATA_START_ROW como fallback
        if max_row < data_start_row:
            max_row = data_start_row
        
        print(
            f"  Procesando: filas {data_start_row} a {max_row} "
            f"({max_row - data_start_row + 1} registros)"
        )
        
        # Iterar solo hasta la última fila con datos
        for row in ws.iter_rows(
            min_row=data_start_row, max_row=max_row, max_col=max_col
        ):
            fila = []
            for cell in row:
                if cell.value is None:
                    fila.append(None)
                    continue
                if cell.is_date:
                    fila.append(cell.value)
                    continue
                if cell.data_type == "n":
                    dec = infer_decimals(cell.number_format)
                    if dec is None:
                        fila.append(cell.value)
                    else:
                        try:
                            fila.append(round(float(cell.value), dec))
                        except Exception:
                            fila.append(cell.value)
                    continue
                fila.append(cell.value)
            data_rows.append(fila)
        
        df = pd.DataFrame(data_rows)
    else:
        df = pd.read_excel(
            ruta_archivo,
            header=None,
            skiprows=data_start_row - 1,
            engine=engine
        )

    # Ajustar encabezados al numero de columnas reales
    if len(df.columns) <= len(encabezados):
        columnas = encabezados[: len(df.columns)]
    else:
        columnas = encabezados
        df = df.iloc[:, : len(encabezados)]
        print(
            f"Aviso: {os.path.basename(ruta_archivo)} tiene {len(df.columns)} columnas, "
            f"encabezados base {len(encabezados)}. Se recortan columnas extra."
        )

    df_salida = pd.DataFrame(df.values, columns=columnas)
    
    # IMPORTANTE: Filtrar solo filas que tienen CONSECUTIVO (primera columna con valor)
    # Esto evita copiar filas vacías o parciales
    primera_columna = columnas[0] if len(columnas) > 0 else None
    if primera_columna:
        # Eliminar filas donde el consecutivo es None, vacío o NaN
        df_salida = df_salida[df_salida[primera_columna].notna()]
        df_salida = df_salida[df_salida[primera_columna].astype(str).str.strip() != '']
    
    num_registros = len(df_salida)
    print(f"    ✓ {num_registros} registros válidos (con consecutivo)")

    # Guardar como .xlsx con sufijo
    base, _ = os.path.splitext(ruta_archivo)
    salida = f"{base}_copia.xlsx"
    df_salida.to_excel(salida, index=False)

    print(f"OK - Generado: {salida}")
    return salida


def generar_con_encabezados(
    archivo_excel,
    encabezados_json_path=ARCHIVO_ENCABEZADOS_JSON,
    data_start_row=DATA_START_ROW,
):
    encabezados = leer_encabezados_json(encabezados_json_path)
    if not encabezados:
        print("No se pudieron leer encabezados.")
        return None
    return procesar_archivo(archivo_excel, encabezados, data_start_row=data_start_row)



def main():
    encabezados = leer_encabezados_json(ARCHIVO_ENCABEZADOS_JSON)
    if not encabezados:
        print("No se pudieron leer encabezados.")
        return

    for root, dirs, files in os.walk(BASE_DIR):
        # Evitar procesar la carpeta scripts_auxiliares y reportes
        if os.path.basename(root).lower() in ["scripts_auxiliares", "reportes", "__pycache__", ".git"]:
            continue

        for nombre in files:
            if nombre.startswith("~$"):
                continue
            if not nombre.lower().endswith(".xlsb"):
                continue
            ruta = os.path.join(root, nombre)
            procesar_archivo(ruta, encabezados, data_start_row=DATA_START_ROW)


if __name__ == "__main__":
    main()
